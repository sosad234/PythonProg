import PyForms
from abc import ABC, abstractmethod
from docx import Document
from openpyxl import Workbook

class BankService(ABC):
    def __init__(self, amount, rate, term):
        self.amount = amount
        self.rate = rate
        self.term = term

    @abstractmethod
    def calculate(self):
        pass

class CreditService(BankService):
    def calculate(self):
        payment = self.amount / self.term
        return f"Ежемесячный платеж: {payment}"

    def __str__(self):
        return "Кредитный сервис"

class InstallmentService(BankService):
    def calculate(self):
        payment = self.amount / self.term
        return f"Ежемесячный платеж: {payment}"

    def __str__(self):
        return "Сервис рассрочки"

class DepositService(BankService):
    def calculate(self):
        return "Вклад не предполагает ежемесячных платежей"

    def __str__(self):
        return "Сервис вкладов"

def save_to_doc(result):
    doc = Document()
    doc.add_paragraph(f'Результат: {result}')
    doc.save('результат.docx')

def save_to_xls(result):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Результат:'
    ws['B1'] = result
    wb.save('результат.xlsx')

app = PyForms.App()

class BankForm(PyForms.Form):
    def __init__(self):
        super().__init__(title='Банковские услуги')

        self.service_type = PyForms.Dropdown('Выберите услугу', ['Кредит', 'Рассрочка', 'Вклад'])
        self.amount = PyForms.FloatInput('Введите сумму', default=0)
        self.rate = PyForms.FloatInput('Введите процентную ставку', default=0)
        self.term = PyForms.IntInput('Введите срок (в месяцах)', default=0)

        self.calculate_button = PyForms.Button('Рассчитать')
        self.save_button = PyForms.Button('Сохранить')
        self.exit_button = PyForms.Button('Выход')

        self.result = PyForms.Label('Результат')

        self.layout = [
            [self.service_type, self.amount, self.rate, self.term],
            [self.calculate_button, self.save_button, self.exit_button],
            [self.result]
        ]

    def button_event(self, event):
        if event == 'Рассчитать':
            if self.service_type.value == 'Кредит':
                service = CreditService(self.amount.value, self.rate.value, self.term.value)
            elif self.service_type.value == 'Рассрочка':
                service = InstallmentService(self.amount.value, self.rate.value, self.term.value)
            elif self.service_type.value == 'Вклад':
                service = DepositService(self.amount.value, self.rate.value, self.term.value)
            result = service.calculate()
            self.result.value = result
        elif event == 'Сохранить':
            if service is not None:
                save_to_doc(result)
                save_to_xls(result)
                self.result.value = 'Результат сохранен.'
            else:
                self.result.value = 'Рассчитайте результаты перед сохранением.'
        elif event == 'Выход':
            app.exit()

app.add_form(BankForm)
app.main()
